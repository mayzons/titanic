import cx_Oracle
import os
import openpyxl
from collections import Counter
from utils.logs_escrita import log_info
from utils.caminhos import caminhos
import sqlite3


def opc_dados(caminho_opc, caminho_arq, nome_arq):
    lista_opc = []

    log_info('Carregando dados do OPC para TRN Domingo')

    # Carrega a Planilha
    workbook = openpyxl.load_workbook(f'{caminho_opc}\\OPC.xlsx')
    sheet = workbook["Page 1"]
    sheet = workbook.active

    # Acessando os dados
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=1, max_col=16): # type: ignore # noqa
        codigo = row[0].value
        servico = row[15].value

        if servico == "Abastece":
            lista_remover = [" - V4 NUC", " - SPP", " - 3"]
            for item in lista_remover:
                codigo = codigo.replace(item, "")  # type: ignore
            lista_opc.append(codigo)  # type: ignore

    workbook.close()

    log_info(f'O OPC foi carregado com sucesso, {len(lista_opc)} postos.')

    return lista_opc


def executa(caminho_arq, nome_arq):
    lista_trn = []
    caminho = caminhos()

    log_info('Iniciado o select no Oracle para TRN Domingo!')

    oracle_instant_client_dir = caminho['ORACLE_HOME']
    # # Adicione o caminho do Oracle Instant Client ao PATH
    os.environ["PATH"] = f"{oracle_instant_client_dir};{os.environ['PATH']}"

    # Verifique se a biblioteca do cliente Oracle já foi inicializada
    if not cx_Oracle.clientversion():
        cx_Oracle.init_oracle_client(lib_dir=oracle_instant_client_dir)

    def conexaoDB():
        conn = cx_Oracle.connect(
            user=caminho['USER'],
            password=caminho['PASSWORD'],
            dsn=f'{caminho['DNS']}:{caminho['PORT']}/{caminho['SERVICE']}'
        )
        return conn

    conexao = conexaoDB()

    resultado = conexao.cursor()

    sql = """
    SELECT TB0008_CD_CONVENIADO AS Postos,
        TO_CHAR(tb0153_dt_transacao, 'YYYY-MM-DD') AS domingo,
        COUNT(*) AS quantidade_registros
    FROM tb0153_transacaoconveniado
    WHERE tb0153_dt_transacao >= TRUNC(SYSDATE, 'IW') - 21
    AND TO_CHAR(
    tb0153_dt_transacao, 'DY', 'NLS_DATE_LANGUAGE=PORTUGUESE') = 'DOM'
    AND tb0138_cd_produto = '1'
    GROUP BY TB0008_CD_CONVENIADO, TO_CHAR(tb0153_dt_transacao, 'YYYY-MM-DD')
    ORDER BY domingo DESC
    """

    resultado.execute(sql)

    for cod, data_tr, qtd in resultado:
        # print(f"Posto: {cod}, Data: {data_tr}, Quantidade: {qtd}")
        lista_trn.append(cod)

    resultado.close()

    log_info(f'As transações foram carregadas com sucesso, {len(lista_trn)} TRN Domingo.')  # noqa

    return lista_trn


def gera_expurgo(list_opc, list_trn, arquivo_abon, data_ontem):

    log_info('Iniciado a geração dos Expurgos para Domingos!')

    expurgo_final = []
    lista_s_trn = []

    # global lista_opc, lista_trn, lista_s_trn, expurgo_final

    lista_opc = [str(c) for c in list_opc]
    lista_trn = [str(c) for c in list_trn]

    contagem_domingo = Counter(lista_trn)

    for codigo_uni in lista_opc:
        quantidade = contagem_domingo.get(codigo_uni, 0)
        if quantidade == 0:
            lista_s_trn.append(int(codigo_uni))

    # Adiciona info
    for i in lista_s_trn:
        expurgo_final.append((i, "Não", data_ontem, "Sim"))
        # print(expurgo_final)

    workbook = openpyxl.load_workbook(
        f'{arquivo_abon}\\S TRN - Abonos Domingos.xlsx')
    sheet = workbook["Abono"]

    l_inicial = sheet.max_row + 1  # type: ignore
    for i, linha in enumerate(expurgo_final, start=l_inicial):
        for j, valor in enumerate(linha, start=1):
            sheet.cell(row=i, column=j, value=valor)  # type: ignore

    workbook.save(f'{arquivo_abon}\\S TRN - Abonos Domingos.xlsx')

    log_info(f'Os abonos foram adicionados com sucesso a planilha, {arquivo_abon}.')  # noqa


def insert_sqlite(exec):
    caminho = caminhos()
    # Criar conexão do SQLite
    log_info('Iniciado a Insersão no banco de dados SQL!')

    connsqlite = sqlite3.connect(caminho['BANCOSQLITE'])
    cursorsqlite = connsqlite.cursor()

    cursorsqlite.execute("""
        CREATE TABLE IF NOT EXISTS execucao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            frente TEXT, Data TEXT, Hora, forma
        );
        """)

    insert_query = """
        INSERT INTO execucao (
            frente, Data, Hora, forma
        ) VALUES (?, ?, ?, ?)
        """

    cursorsqlite.executemany(insert_query, exec)
    connsqlite.commit()
    connsqlite.close()

    log_info(f'Execução {exec} inserida no SQLite com sucesso!')


def data_exec():
    caminho = caminhos()
    log_info('Validando a data de Execução no banco SQL!')

    conexao = sqlite3.connect(caminho['BANCOSQLITE'])
    cursor = conexao.cursor()

    cursor.execute("""
        SELECT max(Data)
        FROM execucao
        where frente = 'trn domingo'
    """)
    rows = cursor.fetchone()[0]

    cursor.close()
    conexao.close()

    return rows
