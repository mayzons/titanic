import cx_Oracle
import sqlite3
import os
from utils.logs_escrita import log_info, log_error
import openpyxl
from openpyxl import Workbook
from utils.caminhos import caminhos


def executa(de, ate):
    expu_autorizacao = []
    caminho = caminhos()

    log_info('Iniciado o select no Oracle para Expur Autorizações!')

    oracle_instant_client_dir = caminho['ORACLE_HOME']
    os.environ["PATH"] = f"{oracle_instant_client_dir};{os.environ['PATH']}"

    if not cx_Oracle.clientversion():
        cx_Oracle.init_oracle_client(lib_dir=oracle_instant_client_dir)

    def conexaoDB():
        conn = cx_Oracle.connect(
            user=caminho['USER'],
            password=caminho['PASSWORD'],
            dsn=f'{caminho['DNS']}:{caminho['PORT']}/{caminho['SERVICE']}'
        )
        return conn

    conexao = conexaoDB()
    resultado = conexao.cursor()

    sql = """
    SELECT
        TB0008_CD_CONVENIADO AS POSTO,
        TO_CHAR(TB0110_DT_PEDIDOAUTORIZACAO, 'YYYY-MM-DD HH24:MI:SS') AS Data,
        TB0138_CD_PRODUTO AS Produto,
        TB0110_ID_DISPOSITIVO as Dispositivo,
        TB0110_NR_PLACAVEICULO as Placa,
        TB0110_VL_SALDODISPONIVEL as Saldo,
        TB0110_FL_RESULTADO AS Resultado,
        TB0025_CL_CDMOTIVOREJEICAO As "Cod. Rejeicao",
        TB0025_CD_MOTIVOREJEICAO As "Mot. Rejeicao",
        TB0110_CD_TOKEN AS TP_SOLICITACAO,
        TB0110_VL_SALDODIARIODISP AS SALDO_DIARIO,
        TB0100_ID_PEDIDO AS ID_PEDIDO,
        TB0110_CD_AUTHEXCEPT AS WL_LIBERADO
    FROM
        CADMO.TB0110_AUTORIZACAO
    WHERE
        TB0110_DT_PEDIDOAUTORIZACAO
                    BETWEEN TO_TIMESTAMP(:data_de, 'DD/MM/YYYY HH24:MI')
                            AND TO_TIMESTAMP(:data_ate, 'DD/MM/YYYY HH24:MI')
        AND TB0025_CD_TIPO = 3
    ORDER BY
        TB0110_DT_PEDIDOAUTORIZACAO DESC
    """

    data_de_str = de
    data_ate_str = ate

    resultado.execute(sql, {'data_de': data_de_str, 'data_ate': data_ate_str})

    rows = resultado.fetchall()

    expu_autorizacao.clear()
    expu_autorizacao.extend(rows)

    resultado.close()
    conexao.close()

    log_info(f'O Select retornou {len(expu_autorizacao)} Autorizações!')

    return expu_autorizacao


def insert_sqlite(exp_aut):
    caminho = caminhos()
    # Criar conexão do SQLite
    resetar_tabela_transacoes()

    log_info('Iniciado a Insersão no banco de dados SQL!')

    connsqlite = sqlite3.connect(caminho['BANCOSQLITE'])
    cursorsqlite = connsqlite.cursor()

    cursorsqlite.execute("""
        CREATE TABLE IF NOT EXISTS expur_autorizacao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            POSTO TEXT, Data TEXT, Produto TEXT,
            Dispositivo TEXT, Placa TEXT, Saldo REAL,
            Resultado TEXT, Cod_Rejeicao TEXT,
            Mot_Rejeicao TEXT, TP_SOLICITACAO TEXT,
            SALDO_DIARIO REAL, ID_PEDIDO TEXT, WL_LIBERADO TEXT
        );
        """)

    insert_query = """
        INSERT INTO expur_autorizacao (
            POSTO, Data, Produto, Dispositivo, Placa, Saldo,
            Resultado, Cod_Rejeicao, Mot_Rejeicao,
            TP_SOLICITACAO, SALDO_DIARIO, ID_PEDIDO, WL_LIBERADO
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

    cursorsqlite.executemany(insert_query, exp_aut)
    connsqlite.commit()
    connsqlite.close()

    log_info(f'{len(exp_aut)} registros inseridos no SQLite com sucesso!')


def resetar_tabela_transacoes():
    caminho = caminhos()

    log_info('Iniciando o reset do banco')

    conexao = sqlite3.connect(caminho['BANCOSQLITE'])
    cursor = conexao.cursor()

    try:
        # Apaga todos os registros
        cursor.execute("DELETE FROM expur_autorizacao;")
        conexao.commit()

        # Compacta o banco (libera espaço)
        cursor.execute("VACUUM;")

        # Reseta o autoincremento (se necessário)
        cursor.execute("""UPDATE sqlite_sequence
                            SET seq = 0
                                WHERE name = 'expur_autorizacao';""")

        conexao.commit()

        log_info('Banco resetado')

    except Exception as e:
        log_error(f'Erro ao resetar a tabela: {e}')

    finally:
        cursor.close()
        conexao.close()


def abrir_arquivo_crit(data_arqui):
    caminho = caminhos()
    log_info('Iniciando a jornada de expurgo com o Critico.xlsx')

    data_str = data_arqui.strftime('%d/%m/%Y')
    novo_nome = data_str.replace('/', '.')

    caminho_arquivo = os.path.join(
        caminho['RP_D_CRITICAL'], f'Critical - {novo_nome}.xlsx')

    workbook = openpyxl.load_workbook(caminho_arquivo)
    sheet = workbook.active

    resultado_final = []

    primeira_linha = True
    for row in sheet.iter_rows(values_only=True):  # type: ignore # noqa
        if primeira_linha:
            cabecalho_original = row
            nova_cabecalho = cabecalho_original + ('Qtd Eventos', 'Possui Evento')  # type: ignore # noqa
            primeira_linha = False
            continue  # pula o cabeçalho

        credenciado = str(row[0])  # type: ignore # noqa
        inicio = row[5]
        final = row[6]

        conexao = sqlite3.connect(caminho['BANCOSQLITE'])
        cursor = conexao.cursor()

        cursor.execute("""
            SELECT COUNT(POSTO) FROM expur_autorizacao
            WHERE Data BETWEEN ? AND ?
            AND POSTO = ?;
        """, (
            inicio.strftime('%Y-%m-%d %H:%M:%S'),  # type: ignore
            final.strftime('%Y-%m-%d %H:%M:%S'),  # type: ignore
            credenciado
        ))

        qtd_eventos = cursor.fetchone()[0]
        possui_evento = 'Sim' if qtd_eventos > 0 else 'Não'

        linha_formatada = list(row)
        linha_formatada[5] = inicio.strftime('%d/%m/%Y %H:%M')  # type: ignore
        linha_formatada[6] = final.strftime('%d/%m/%Y %H:%M')  # type: ignore

        nova_linha = tuple(linha_formatada + [qtd_eventos, possui_evento])
        resultado_final.append(nova_linha)

        cursor.close()
        conexao.close()

    # Criar e salvar novo Excel
    wb_novo = Workbook()
    ws_novo = wb_novo.active
    ws_novo.append(nova_cabecalho)  # type: ignore  # noqa

    for linha in resultado_final:
        ws_novo.append(linha)  # type: ignore

    data_formatada = data_arqui
    caminho_salvar = os.path.join(
        caminho['RP_AUTORIZA_ABONO'],
        f'Abono - {data_formatada.strftime("%d_%m_%Y")}.xlsx')
    wb_novo.save(caminho_salvar)

    log_info(f'Foram Expurgados {len(resultado_final)} eventos!')
