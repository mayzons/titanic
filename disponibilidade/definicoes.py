import os
import openpyxl  # type: ignore
from openpyxl import Workbook
import csv
from utils.logs_escrita import log_info
from utils.caminhos import caminhos


def nome_col(aq_origem, p_origem, nm_arquivo):
    caminho = caminhos()

    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(aq_origem)
    sheet = workbook.active

    # Inserir uma linha vazia na primeira posição
    sheet.delete_rows(1)  # type: ignore
    sheet.delete_cols(10)  # type: ignore
    # Salvar as alterações no arquivo

    colunas_para_alterar = {
            "Código": "Código",
            "Nome": "Posto",
            "Tecnologia": "Tecnologia",
            "Aplicação": "Aplicação",
            "Tipo": "Componente",
            "Inicio": "Inicio",
            "Fim": "Final",
            "Duração": "Duração",
            "%": "%"
    }

    # Alterar os nomes das colunas
    for col in sheet.iter_cols(min_row=1, max_row=1):  # type: ignore # noqa
        if col[0].value in colunas_para_alterar:  # noqa
            col[0].value = colunas_para_alterar[col[0].value]  # type: ignore

    numero_linhas = int(sheet.max_row)   # type: ignore
    for col2 in sheet.iter_rows(min_row=2,max_row=numero_linhas):  # type: ignore # noqa
        col2[0].value = int(col2[0].value)  # type: ignore # noqa
    planilha = workbook['Sheet1']
    planilha.title = 'Plan1'
    caminho_destino_salvar = os.path.join(
        caminho['RP_D_CRITICAL'], os.path.basename(nm_arquivo))
    workbook.save(caminho_destino_salvar)
    workbook.close()
    os.remove(aq_origem)

    log_info(f'Arquivo Tratado {nm_arquivo}')


def apaga_linha(aq_origem, p_origem, nm_arquivo):
    caminho = caminhos()
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(aq_origem)
    sheet = workbook.active
    # Excluir as duas primeiras linhas
    sheet.delete_rows(1)  # type: ignore

    if 'sla_antena_geral_abastece' in aq_origem.lower():
        caminho_seguir = caminho['RP_ANTENAS']

        log_info(f'Arquivo Tratado {nm_arquivo}')

    elif 'sla_abastece' in aq_origem.lower():
        caminho_seguir = caminho['RP_SLA']
        log_info(f'Arquivo Tratado {nm_arquivo}')

    else:
        os.remove(aq_origem)
        log_info(f'Arquivo removido {nm_arquivo}')

    # Salvar o arquivo na pasta de destino
    caminho_destino = os.path.join(
        caminho_seguir, os.path.basename(aq_origem))
    workbook.save(caminho_destino)
    # Fechar o workbook
    workbook.close()
    # Apagar o arquivo original na pasta de origem
    os.remove(aq_origem)


def gera_disp(aq_origem, p_origem, nm_arquivo):
    caminho = caminhos()

    # Nome dos arquivos
    nome_csv, extensao_csv = os.path.splitext(nm_arquivo)
    nome_excel = f'{nome_csv}.xlsx'

    # Criar uma nova planilha do Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"  # type: ignore

    # Ler o arquivo CSV e escrever no Excel
    with open(aq_origem, newline='', encoding='utf-8') as f:
        leitor_csv = csv.reader(f)  # type: ignore  # noqa
        for linha in leitor_csv:
            ws.append(linha)  # type: ignore

    # Salvar o arquivo Excel
    wb.save(os.path.join(
        caminho['RP_D_PROC'], nome_excel))

    log_info(f'Arquivo tratado {nm_arquivo}')
    os.system(f'copy "{aq_origem}" "{caminho['RP_D_PROC_ORI']}"')

    log_info(f'Arquivo salvo {nm_arquivo}')
    os.remove(aq_origem)
